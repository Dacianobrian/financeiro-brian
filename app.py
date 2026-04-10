import re
import io
import os
import hashlib
import urllib.request
from datetime import date, datetime
from functools import wraps
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for

import database as db

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'brian-financeiro-2026-local')

# ── Autenticação simples ──────────────────────────────────────────────────────
APP_PASSWORD = os.environ.get('APP_PASSWORD', 'brian2026')

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        senha = request.form.get('senha', '')
        if senha == APP_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('index'))
        error = 'Senha incorreta.'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ── Investidor 10 scraper ──────────────────────────────────────────────────────

def _fetch_investidor10(ticker: str) -> dict:
    """Fetch price, DY, P/VP and last dividend from Investidor 10."""
    url = f"https://investidor10.com.br/fiis/{ticker.lower()}/"
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/123.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
        "Accept-Language": "pt-BR,pt;q=0.9",
        "Accept-Encoding": "identity",
    })
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            html = resp.read().decode("utf-8", errors="ignore")
    except Exception as e:
        return {"error": str(e)}

    def _br(s):
        s = re.sub(r'[R$\s%]', '', s).strip()
        if ',' in s and '.' in s:
            s = s.replace('.', '').replace(',', '.')
        elif ',' in s:
            s = s.replace(',', '.')
        try:
            return float(s)
        except Exception:
            return None

    result = {}

    # ── Cotação: <span title="Cotação">...</span> ... <span class="value">R$ 9,74</span>
    m = re.search(
        r'<span[^>]*title="Cota[çc][ãa]o"[^>]*>[\s\S]{0,400}?'
        r'<span[^>]*class="[^"]*value[^"]*"[^>]*>\s*(R\$[\s\d,\.]+)',
        html, re.DOTALL | re.IGNORECASE
    )
    if m:
        result["price"] = _br(m.group(1))

    # ── P/VP: _card com "P/VP" → value
    m = re.search(
        r'<span[^>]*title="P/VP"[^>]*>[\s\S]{0,400}?'
        r'<span[^>]*class="[^"]*value[^"]*"[^>]*>\s*([\d,\.]+)',
        html, re.DOTALL | re.IGNORECASE
    )
    if m:
        result["pvp"] = _br(m.group(1))

    # ── DY 12M: texto FAQ "Dividend Yield no período foi de 12,26%"
    m = re.search(r'Dividend\s+Yield[^d]{0,50}?de\s+([\d,\.]+)\s*%', html, re.IGNORECASE | re.DOTALL)
    if not m:
        m = re.search(r'Dividend\s+Yeld[^d]{0,50}?([\d,\.]+)\s*%', html, re.IGNORECASE | re.DOTALL)
    if m:
        result["dy"] = _br(m.group(1))

    # ── Média mensal de dividendos: "média mensal de R$ 0,10"
    m = re.search(r'm[eé]dia\s+mensal\s+de\s+R\$\s*([\d,\.]+)', html, re.IGNORECASE | re.DOTALL)
    if m:
        result["div"] = _br(m.group(1))

    return result


def atualizar_fiis_investidor10() -> tuple[bool, str]:
    """Fetch live quotes for all tickers in portfolio and update DB."""
    conn = db.get_conn()
    c = conn.cursor()
    tickers = [r[0] for r in c.execute("SELECT ticker FROM fii_portfolio ORDER BY ticker").fetchall()]
    conn.close()

    if not tickers:
        return False, "Nenhum FII na carteira."

    lines = []
    erros = []
    for ticker in tickers:
        data = _fetch_investidor10(ticker)
        if "error" in data or "price" not in data:
            erros.append(ticker)
            lines.append(f"  ✗ {ticker}: falha na busca")
            continue

        price = data.get("price")
        dy    = data.get("dy")
        pvp   = data.get("pvp")
        div   = data.get("div")

        conn = db.get_conn()
        cc = conn.cursor()
        cc.execute("""UPDATE fii_portfolio
                      SET price=?,
                          dy   = COALESCE(?, dy),
                          pvp  = COALESCE(?, pvp),
                          dividends_per_share = COALESCE(?, dividends_per_share)
                      WHERE ticker=?""",
                   (price, dy, pvp, div, ticker))
        conn.commit()
        conn.close()

        div_str = f"div R${div:.3f}" if div else ""
        dy_str  = f"DY {dy:.1f}%" if dy else ""
        lines.append(f"  ✓ {ticker}: R$ {price:.2f}  {dy_str}  {div_str}")

    summary = f"Cotações atualizadas ({len(tickers)-len(erros)}/{len(tickers)} FIIs):\n"
    summary += "\n".join(lines)
    if erros:
        summary += f"\n  ⚠ Erros: {', '.join(erros)}"
    return True, summary

# ── Utility ───────────────────────────────────────────────────────────────────

def parse_br_value(raw: str) -> float:
    """Accept R$5.000,00 / 5.000 / 5000,50 / 5000.50 etc."""
    s = raw.strip()
    s = re.sub(r'[Rr]\$\s*', '', s)  # remove R$
    s = s.strip()
    # If there's a comma and a dot, the format is 1.234,56
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        # Could be 1234,56 (decimal comma) or 1.234 (thousands dot only)
        # If comma is the last separator and 2 digits after, treat as decimal
        parts = s.split(',')
        if len(parts) == 2 and len(parts[1]) <= 2:
            s = s.replace(',', '.')
        else:
            s = s.replace(',', '')
    else:
        # Remove thousand-dots: if there's a dot not at decimal position
        # e.g. 5.000  vs  5.50
        dot_parts = s.split('.')
        if len(dot_parts) > 1 and len(dot_parts[-1]) == 3:
            s = s.replace('.', '')
    return float(s)


MONTH_MAP = {
    'jan': 'Janeiro', 'fev': 'Fevereiro', 'mar': 'Março',
    'abr': 'Abril',  'mai': 'Maio',      'jun': 'Junho',
    'jul': 'Julho',  'ago': 'Agosto',    'set': 'Setembro',
    'out': 'Outubro','nov': 'Novembro',  'dez': 'Dezembro',
    'janeiro': 'Janeiro', 'fevereiro': 'Fevereiro', 'março': 'Março',
    'abril': 'Abril', 'maio': 'Maio', 'junho': 'Junho',
    'julho': 'Julho', 'agosto': 'Agosto', 'setembro': 'Setembro',
    'outubro': 'Outubro', 'novembro': 'Novembro', 'dezembro': 'Dezembro',
}

HELP_TEXT = """── FORMATO PADRÃO ──────────────────────────────────────────────────────────
  DD/MM/AAAA - valor conta - descrição          → saída (padrão)
  entrada  DD/MM/AAAA - valor conta - descrição → receita
  investido DD/MM/AAAA - valor conta - descrição→ investimento

  Exemplos:
    08/04/2026 - 5,00 corrente - compra no mercado
    08/04/2026 - 150,00 cartão bradesco - almoço restaurante
    entrada 08/04/2026 - 3.500,00 nubank - comissão imobiliária
    investido 08/04/2026 - 500,00 nubank - aporte FII

── OUTROS COMANDOS ──────────────────────────────────────────────────────────
  nf [ref] [valor]                   — Nota fiscal MEI
  das [mês] [valor]                  — DAS MEI (ex: das abr 81.50)
  fii [ticker] [cotas] [preço]       — Atualiza FII
  fii [ticker] preço [valor]         — Atualiza apenas preço
  fii [ticker] +[cotas] [preço]      — Adiciona cotas
  atualizar fiis                     — Busca cotações ao vivo
  conta [nubank|bradesco] [valor]    — Atualiza saldo de conta
  caixinha [valor]                   — Atualiza saldo da caixinha
  listar / lista                     — Últimas 10 transações
  saldo                              — Saldo do mês atual
  ajuda / help                       — Esta mensagem"""


# ── Auto-categorização por palavras-chave ─────────────────────────────────────
def auto_category(desc: str, tipo: str) -> str:
    d = desc.lower()
    if tipo == 'entrada':
        if any(w in d for w in ('comissão', 'comissao', 'imobil')):
            return 'Comissão Imobiliária'
        if any(w in d for w in ('salário', 'salario', 'pgto', 'pagamento')):
            return 'Salário'
        if any(w in d for w in ('dividendo', 'fii', 'rend')):
            return 'Dividendos FII'
        if any(w in d for w in ('rdb', 'cdb', 'caixinha', 'juros')):
            return 'Rendimento RDB/CDB'
        if any(w in d for w in ('mei', 'nota', 'nf', 'cnpj')):
            return 'Gerenciamento'
        return 'Entrada'
    if tipo == 'investido':
        return 'Investimentos'
    # saída
    if any(w in d for w in ('mercado', 'supermercado', 'feira', 'hortifruti', 'atacadão', 'atacado', 'atacarejo')):
        return 'Supermercado'
    if any(w in d for w in ('restaurante', 'almoço', 'almoco', 'jantar', 'lanche', 'pizza', 'hamburguer',
                             'hamburger', 'sushi', 'churrasco', 'padaria', 'café', 'cafe', 'bakery')):
        return 'Alimentação'
    if any(w in d for w in ('posto', 'gasolina', 'combustivel', 'combustível', 'etanol', 'diesel')):
        return 'Combustível'
    if any(w in d for w in ('uber', '99', 'taxi', 'táxi', 'onibus', 'ônibus', 'passagem', 'metrô', 'metro')):
        return 'Transporte'
    if any(w in d for w in ('farmacia', 'farmácia', 'remedio', 'remédio', 'medico', 'médico',
                             'consulta', 'exame', 'clinica', 'clínica', 'hospital', 'drogaria')):
        return 'Saúde'
    if any(w in d for w in ('academia', 'gym', 'pilates', 'yoga')):
        return 'Saúde'
    if any(w in d for w in ('netflix', 'spotify', 'amazon', 'disney', 'hbo', 'apple', 'youtube',
                             'streaming', 'prime')):
        return 'Streaming'
    if any(w in d for w in ('aluguel', 'locação', 'locacao', 'locabens')):
        return 'Moradia'
    if any(w in d for w in ('energia', 'cpfl', 'luz', 'conta de luz')):
        return 'Moradia'
    if any(w in d for w in ('água', 'agua', 'semae', 'saae', 'corsan')):
        return 'Moradia'
    if any(w in d for w in ('internet', 'fibra', 'vivo', 'claro', 'tim', 'oi', 'telecom')):
        return 'Moradia'
    if any(w in d for w in ('cartão', 'cartao', 'fatura', 'bradesco')):
        return 'Cartão Bradesco'
    if any(w in d for w in ('das', 'mei', 'imposto', 'ir', 'irpf')):
        return 'Impostos / MEI'
    return 'Outros'


def parse_command(cmd: str):
    """Parse a Portuguese command string. Returns (success, message, data)."""
    cmd = cmd.strip()
    if not cmd:
        return False, "Comando vazio.", {}

    parts = cmd.split()
    verb = parts[0].lower()
    today_str = date.today().isoformat()
    current_year = date.today().year

    # ── FORMATO PADRÃO: [tipo] DD/MM/AAAA - valor [conta] - descrição ────────
    # Detecta se o comando segue o padrão data-valor-descrição
    _std = re.match(
        r'^(?:(entrada|saida|saída|investido|investimento)\s+)?'   # tipo (opcional)
        r'(\d{2}/\d{2}/\d{4})\s*[-–]\s*'                          # data
        r'([\d.,]+)\s*'                                             # valor
        r'([^-–]*?)\s*[-–]\s*'                                     # conta/cartão (opcional)
        r'(.+)$',                                                   # descrição
        cmd, re.IGNORECASE
    )
    if _std:
        _tipo_raw  = (_std.group(1) or 'saida').lower()
        _date_raw  = _std.group(2)         # DD/MM/AAAA
        _val_raw   = _std.group(3)
        _conta_raw = (_std.group(4) or '').strip()
        _desc_raw  = _std.group(5).strip()

        # Normaliza tipo
        tipo = 'investido' if _tipo_raw in ('investido','investimento') \
               else 'entrada' if _tipo_raw == 'entrada' \
               else 'saida'

        # Converte data DD/MM/AAAA → AAAA-MM-DD
        try:
            _d, _m, _y = _date_raw.split('/')
            date_iso = f"{_y}-{_m.zfill(2)}-{_d.zfill(2)}"
            datetime.strptime(date_iso, "%Y-%m-%d")   # valida
        except Exception:
            return False, f"Data inválida: {_date_raw}. Use DD/MM/AAAA.", {}

        try:
            amount = parse_br_value(_val_raw)
        except Exception:
            return False, f"Valor inválido: {_val_raw}", {}

        # Monta descrição incluindo conta se informada
        conta_label = _conta_raw.strip()
        if conta_label:
            # Detecta banco para badge correto
            if re.search(r'bradesco', conta_label, re.I):
                desc = f"{_desc_raw} [Bradesco]"
            elif re.search(r'nubank|nu', conta_label, re.I):
                desc = f"{_desc_raw} [Nubank]"
            else:
                desc = f"{_desc_raw} [{conta_label}]"
        else:
            desc = _desc_raw

        category = auto_category(_desc_raw, tipo)
        db.add_transaction(date_iso, tipo, amount, desc, category)

        tipo_label = '💵 Entrada' if tipo=='entrada' else ('📈 Investimento' if tipo=='investido' else '📤 Saída')
        return True, (
            f"{tipo_label} registrada!\n"
            f"  Data: {_date_raw}  ·  Valor: R$ {amount:,.2f}\n"
            f"  Conta: {conta_label or '—'}  ·  Categoria: {category}\n"
            f"  Descrição: {_desc_raw}"
        ), {}

    # ── ajuda / help ──────────────────────────────────────────────────────────
    if verb in ('ajuda', 'help'):
        return True, HELP_TEXT, {}

    # ── listar / lista ────────────────────────────────────────────────────────
    if verb in ('listar', 'lista'):
        rows = db.get_recent_transactions(10)
        return True, f"{len(rows)} transações recentes carregadas.", {"transactions": rows}

    # ── saldo ─────────────────────────────────────────────────────────────────
    if verb == 'saldo':
        d = db.get_current_month_data()
        msg = (
            f"Mês atual — Entradas: R$ {d['entradas']:,.2f} | "
            f"Saídas: R$ {d['saidas']:,.2f} | "
            f"Investido: R$ {d['investido']:,.2f} | "
            f"Saldo: R$ {d['saldo']:,.2f}"
        )
        return True, msg, d

    # ── entrada ───────────────────────────────────────────────────────────────
    if verb == 'entrada':
        if len(parts) < 2:
            return False, "Uso: entrada [valor] [descrição]", {}
        try:
            amount = parse_br_value(parts[1])
        except Exception:
            return False, f"Valor inválido: {parts[1]}", {}
        desc = ' '.join(parts[2:]) if len(parts) > 2 else 'Entrada'
        db.add_transaction(today_str, 'entrada', amount, desc)
        return True, f"Entrada de R$ {amount:,.2f} registrada: {desc}", {}

    # ── saída / saida ─────────────────────────────────────────────────────────
    if verb in ('saída', 'saida', 'saida'):
        if len(parts) < 2:
            return False, "Uso: saída [valor] [descrição]", {}
        try:
            amount = parse_br_value(parts[1])
        except Exception:
            return False, f"Valor inválido: {parts[1]}", {}
        desc = ' '.join(parts[2:]) if len(parts) > 2 else 'Saída'
        db.add_transaction(today_str, 'saida', amount, desc)
        return True, f"Saída de R$ {amount:,.2f} registrada: {desc}", {}

    # ── investido / investimento ───────────────────────────────────────────────
    if verb in ('investido', 'investimento'):
        if len(parts) < 2:
            return False, "Uso: investido [valor] [descrição]", {}
        try:
            amount = parse_br_value(parts[1])
        except Exception:
            return False, f"Valor inválido: {parts[1]}", {}
        desc = ' '.join(parts[2:]) if len(parts) > 2 else 'Investimento'
        db.add_transaction(today_str, 'investido', amount, desc)
        return True, f"Investimento de R$ {amount:,.2f} registrado: {desc}", {}

    # ── nf / nota ─────────────────────────────────────────────────────────────
    if verb in ('nf', 'nota'):
        if len(parts) < 3:
            return False, "Uso: nf [ref] [valor]", {}
        ref = parts[1]
        try:
            amount = parse_br_value(parts[2])
        except Exception:
            return False, f"Valor inválido: {parts[2]}", {}
        db.add_mei_invoice(ref, amount, current_year, today_str)
        return True, f"Nota Fiscal {ref} de R$ {amount:,.2f} registrada.", {}

    # ── das ───────────────────────────────────────────────────────────────────
    if verb == 'das':
        if len(parts) < 3:
            return False, "Uso: das [mês] [valor]  (ex: das jan 81.50)", {}
        month_raw = parts[1].lower()
        month_name = MONTH_MAP.get(month_raw)
        if not month_name:
            return False, f"Mês não reconhecido: {parts[1]}", {}
        try:
            amount = parse_br_value(parts[2])
        except Exception:
            return False, f"Valor inválido: {parts[2]}", {}
        db.add_das_payment(month_name, current_year, amount, today_str)
        return True, f"DAS {month_name}/{current_year} de R$ {amount:,.2f} registrado.", {}

    # ── fii ───────────────────────────────────────────────────────────────────
    if verb == 'fii':
        if len(parts) < 3:
            return False, "Uso: fii [ticker] [cotas] [preço]  ou  fii [ticker] preço [valor]", {}
        ticker = parts[1].upper()

        # fii TICKER preço VALOR
        if parts[2].lower() == 'preço' or parts[2].lower() == 'preco':
            if len(parts) < 4:
                return False, "Uso: fii [ticker] preço [valor]", {}
            try:
                price = parse_br_value(parts[3])
            except Exception:
                return False, f"Preço inválido: {parts[3]}", {}
            db.update_fii_price(ticker, price)
            return True, f"Preço de {ticker} atualizado para R$ {price:,.2f}.", {}

        # fii TICKER +COTAS PREÇO
        if parts[2].startswith('+'):
            try:
                extra = int(parts[2][1:])
            except Exception:
                return False, f"Cotas inválidas: {parts[2]}", {}
            price = 0.0
            if len(parts) >= 4:
                try:
                    price = parse_br_value(parts[3])
                except Exception:
                    pass
            db.add_fii_shares(ticker, extra, price)
            return True, f"+{extra} cotas de {ticker} adicionadas ao preço R$ {price:,.2f}.", {}

        # fii TICKER COTAS PREÇO
        try:
            shares = int(parts[2])
        except Exception:
            return False, f"Cotas inválidas: {parts[2]}", {}
        price = 0.0
        if len(parts) >= 4:
            try:
                price = parse_br_value(parts[3])
            except Exception:
                pass
        db.upsert_fii(ticker, shares, price)
        return True, f"FII {ticker}: {shares} cotas @ R$ {price:,.2f} atualizado.", {}

    # ── conta / caixinha ─────────────────────────────────────────────────────
    if verb in ('conta', 'caixinha', 'saldo_conta'):
        # conta nubank 724.60  |  conta bradesco 5150.14  |  caixinha 35275.25
        if verb == 'caixinha':
            name = 'caixinha'
            val_str = parts[1] if len(parts) > 1 else ''
        else:
            if len(parts) < 3:
                return False, "Uso: conta [nubank|bradesco|caixinha] [valor]", {}
            name = ' '.join(parts[1:-1])
            val_str = parts[-1]
        try:
            balance = parse_br_value(val_str)
        except Exception:
            return False, f"Valor inválido: {val_str}", {}
        ok = db.update_account(name, balance)
        if ok:
            return True, f"Conta '{name}' atualizada para R$ {balance:,.2f}.", {}
        return False, f"Conta '{name}' não encontrada. Contas: nubank, bradesco, caixinha.", {}

    # ── atualizar fiis ────────────────────────────────────────────────────────
    if verb in ('atualizar') and len(parts) > 1 and parts[1].lower() == 'fiis':
        ok, msg = atualizar_fiis_investidor10()
        return ok, msg, {}

    return False, f"Comando não reconhecido: '{verb}'. Digite 'ajuda' para ver os comandos.", {}


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    return render_template('index.html')


@app.route('/api/command', methods=['POST'])
@login_required
def api_command():
    body = request.get_json(silent=True) or {}
    cmd = body.get('command', '').strip()
    success, message, data = parse_command(cmd)
    return jsonify({"success": success, "message": message, "data": data})


@app.route('/api/dashboard')
@login_required
def api_dashboard():
    monthly_2026 = db.get_monthly_summary(2026)
    monthly_2025 = db.get_monthly_summary_2025()
    current      = db.get_current_month_data()
    fiis         = db.get_fii_portfolio()

    total_fii         = sum(f['shares'] * f['price'] for f in fiis)
    monthly_dividends = sum(f['shares'] * (f['dividends_per_share'] or 0) for f in fiis)

    rate_2026, inv_2026 = db.get_savings_rate(2026)
    rate_2025, inv_2025 = db.get_savings_rate(2025)
    expense_by_cat      = db.get_expense_by_category(2026)

    accounts   = db.get_accounts()
    patrimonio = db.get_patrimonio_breakdown()

    # Saldo em contas correntes (Nubank conta + Bradesco corrente)
    saldo_contas = sum(
        a['balance'] for a in accounts
        if a['type'].lower() == 'corrente'
    )
    # Total investido real = caixinha (RDB) + FIIs
    caixinha_val = sum(
        a['balance'] for a in accounts
        if 'caixinha' in a['name'].lower() or a['type'].lower() in ('caixinha', 'rdb', 'poupança')
    )
    total_investido_real = caixinha_val + total_fii

    META_RESERVA = 60000.0

    return jsonify({
        "current_month":        current,
        "monthly_2026":         monthly_2026,
        "monthly_2025":         monthly_2025,
        "expense_by_cat":       expense_by_cat,
        "transactions_recent":  db.get_recent_transactions(10),
        "fii":                  fiis,
        "fii_total":            total_fii,
        "fii_monthly_dividends":monthly_dividends,
        "mei_invoices":         db.get_mei_invoices(10),
        "das_payments":         db.get_das_payments(),
        "fixed_costs":          db.get_fixed_costs(),
        "patrimonio_total":     db.get_patrimonio_total(),
        "savings_rate_2026":    rate_2026,
        "total_investido_2026": inv_2026,
        "savings_rate_2025":    rate_2025,
        "total_investido_2025": inv_2025,
        "accounts":             accounts,
        "patrimonio":           patrimonio,
        "saldo_contas":         saldo_contas,
        "total_investido_real": total_investido_real,
        "meta_reserva":         META_RESERVA,
    })


@app.route('/api/export')
@login_required
def api_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="1a1d2e")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    ALT_FILL    = PatternFill("solid", fgColor="F5F5F5")

    def style_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    def fmt_brl(v):
        if v is None:
            return ''
        return f"R$ {v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

    # Sheet 1: Transações
    ws1 = wb.active
    ws1.title = "Transações"
    style_header(ws1, ["ID", "Data", "Tipo", "Valor", "Descrição", "Categoria", "Criado em"])
    for i, t in enumerate(db.get_all_transactions(), 2):
        ws1.append([
            t['id'], t['date'], t['type'],
            fmt_brl(t['amount']), t['description'], t['category'], t['created_at']
        ])
        if i % 2 == 0:
            for cell in ws1[i]:
                cell.fill = ALT_FILL
    auto_width(ws1)

    # Sheet 2: MEI Notas
    ws2 = wb.create_sheet("MEI Notas")
    style_header(ws2, ["ID", "Referência", "Data Emissão", "Valor", "Ano"])
    for i, n in enumerate(db.get_mei_invoices(9999), 2):
        ws2.append([n['id'], n['ref'], n['issue_date'], fmt_brl(n['amount']), n['year']])
        if i % 2 == 0:
            for cell in ws2[i]:
                cell.fill = ALT_FILL
    auto_width(ws2)

    # Sheet 3: DAS
    ws3 = wb.create_sheet("DAS")
    style_header(ws3, ["ID", "Mês", "Ano", "Valor", "Data Pagamento"])
    for i, d in enumerate(db.get_das_payments(), 2):
        ws3.append([d['id'], d['month_name'], d['year'], fmt_brl(d['amount']), d['payment_date']])
        if i % 2 == 0:
            for cell in ws3[i]:
                cell.fill = ALT_FILL
    auto_width(ws3)

    # Sheet 4: FII Portfolio
    ws4 = wb.create_sheet("FII Portfolio")
    style_header(ws4, [
        "Ticker", "Cotas", "Preço", "Valor Total", "Setor", "Tipo",
        "P/VP", "DY %", "Var 12m %", "Div/Cota", "Vacância %"
    ])
    for i, f in enumerate(db.get_fii_portfolio(), 2):
        total = (f['shares'] or 0) * (f['price'] or 0)
        ws4.append([
            f['ticker'], f['shares'], fmt_brl(f['price']), fmt_brl(total),
            f['sector'], f['fund_type'],
            f['pvp'], f['dy'], f['variation_12m'],
            fmt_brl(f['dividends_per_share']), f['vacancy']
        ])
        if i % 2 == 0:
            for cell in ws4[i]:
                cell.fill = ALT_FILL
    auto_width(ws4)

    # Sheet 5: Custos Fixos
    ws5 = wb.create_sheet("Custos Fixos")
    style_header(ws5, ["ID", "Descrição", "Valor", "Dia Venc.", "Tipo"])
    for i, c in enumerate(db.get_fixed_costs(), 2):
        ws5.append([c['id'], c['description'], fmt_brl(c['amount']), c['due_day'], c['type']])
        if i % 2 == 0:
            for cell in ws5[i]:
                cell.fill = ALT_FILL
    auto_width(ws5)

    # Sheet 6: Resumo Mensal
    ws6 = wb.create_sheet("Resumo Mensal")
    style_header(ws6, ["Mês", "Entradas", "Saídas", "Investido", "Saldo"])
    monthly = db.get_monthly_summary_2025()
    for i, m in enumerate(monthly, 2):
        saldo = m['entradas'] - m['saidas'] - m['investido']
        ws6.append([
            m['month'], fmt_brl(m['entradas']),
            fmt_brl(m['saidas']), fmt_brl(m['investido']), fmt_brl(saldo)
        ])
        if i % 2 == 0:
            for cell in ws6[i]:
                cell.fill = ALT_FILL
    auto_width(ws6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"financeiro_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# Inicializa DB ao carregar o módulo (necessário para gunicorn em produção)
db.init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    print()
    print("=" * 55)
    print("  💰  Brian Controle Financeiro")
    print(f"  Acesse: http://127.0.0.1:{port}")
    print("  Pressione CTRL+C para encerrar")
    print("=" * 55)
    print()
    app.run(debug=False, port=port, host='0.0.0.0')
